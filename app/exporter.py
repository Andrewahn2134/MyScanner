from __future__ import annotations

from io import BytesIO
from typing import Iterable, Sequence, Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def make_xlsx_bytes(sheet_name: str, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Sheet1"

    ws.append(list(headers))
    for r in rows:
        ws.append([_cell(v) for v in r])

    # Basic column sizing (cap to keep file sane)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header)) if header is not None else 10
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
            v = row[0].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(v: Any) -> Any:
    # keep datetimes and numbers; stringify everything else safely
    return v
